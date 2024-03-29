import networkx as nx
import numpy as np
from openpyxl import load_workbook
import os


class NodeConnectionGraph(object):
    def __init__(self, node_num, p, k=4, m=5, graph_mode="NCNN", graph_type="ncnn16.xlsx"):
        self.node_num = node_num
        self.p = p
        self.k = k
        self.m = m
        self.graph_mode = graph_mode
        self.graph_type = graph_type

    def xlsx_numpy_matrix(self, filename):
        #load excel file
        workbook = load_workbook(filename=filename)
        worksheet = workbook["Sheet1"]

        rows = worksheet.max_row
        columns = worksheet.max_column
        data = np.zeros((rows, columns))
        index = 0
        for col in worksheet.columns:
            tmp = list()
            for x in col:
                tmp.append(x.value)
            data[:][index] = tmp
            index += 1
        return nx.from_numpy_matrix(data)

    def make_graph(self):
        # reference
        # https://networkx.github.io/documentation/networkx-1.9/reference/generators.html

        # Code details,
        # In the case of the nx.random_graphs module, we can give the random seeds as a parameter.
        # But I have implemented it to handle it in the module.
        if self.graph_mode == "ER":
            graph = nx.random_graphs.erdos_renyi_graph(self.node_num, self.p)
        elif self.graph_mode == "WS":
            graph = nx.random_graphs.connected_watts_strogatz_graph(self.node_num, self.k, self.p)
        elif self.graph_mode == "BA":
            graph = nx.random_graphs.barabasi_albert_graph(self.node_num, self.m)
        elif self.graph_mode == "NCNN":
            graph = self.xlsx_numpy_matrix("./data/" + self.graph_type)
        return graph

    def get_graph_info(self, graph):
        in_edges = {}
        in_edges[0] = []
        nodes = [0]
        end = []
        for node in graph.nodes():
            neighbors = list(graph.neighbors(node))
            neighbors.sort()

            edges = []
            check = []
            for neighbor in neighbors:
                if node > neighbor:
                    edges.append(neighbor + 1)
                    check.append(neighbor)
            if not edges:
                edges.append(0)
            in_edges[node + 1] = edges
            if check == neighbors:
                end.append(node + 1)
            nodes.append(node + 1)
        in_edges[self.node_num + 1] = end
        nodes.append(self.node_num + 1)

        return nodes, in_edges

    def save_random_graph(self, graph, path):
        if not os.path.isdir("saved_graph"):
            os.mkdir("saved_graph")
        nx.write_yaml(graph, "./saved_graph/" + path)

    def load_random_graph(self, path):
        return nx.read_yaml("./saved_graph/" + path)
